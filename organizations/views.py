from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.http import JsonResponse
from django.db.models import Prefetch
from .models import Padaria, PadariaUser, ApiKey, Promocao, Produto, Cliente, CampanhaWhatsApp, MensagemCampanha
from audit.models import AuditLog
import requests


def get_user_padarias(user):
    """Retorna padarias que o usuário pode acessar."""
    if user.is_superuser:
        return Padaria.objects.all()
    user_padaria_ids = PadariaUser.objects.filter(user=user).values_list('padaria_id', flat=True)
    return Padaria.objects.filter(id__in=user_padaria_ids)


def send_products_webhook(padaria, user=None, action="product_update"):
    """
    Envia todos os produtos da padaria para o webhook N8N para atualização do RAG.
    Formata produtos como catálogo de texto incluindo promoções ativas.
    """
    try:
        from django.utils import timezone
        from django.utils.text import slugify
        
        # Buscar agente da padaria
        agent = padaria.get_agent()
        if not agent:
            print(f"[DEBUG] Padaria {padaria.name} não tem agente. Pulando webhook de produtos.")
            return False
        
        # Buscar todos os produtos ativos da padaria
        produtos = Produto.objects.filter(padaria=padaria, ativo=True).order_by('categoria', 'nome')
        
        if not produtos.exists():
            print(f"[DEBUG] Padaria {padaria.name} não tem produtos ativos. Pulando webhook.")
            return False
        
        # Buscar promoções ativas vinculadas a produtos
        promocoes_ativas = Promocao.objects.filter(
            padaria=padaria,
            produto__isnull=False,
            is_active=True
        ).select_related('produto')
        
        # Criar dicionário de produto_id -> promoção ativa
        promocao_por_produto = {}
        for promo in promocoes_ativas:
            if promo.is_valid() and promo.produto_id:
                promocao_por_produto[promo.produto_id] = promo
        
        # Formatar produtos como catálogo de texto
        linhas = ["Cardápio / Catálogo de Produtos", ""]
        
        categoria_atual = None
        for produto in produtos:
            # Adicionar cabeçalho de categoria se mudou
            if produto.categoria and produto.categoria != categoria_atual:
                if categoria_atual is not None:
                    linhas.append("")  # Linha em branco entre categorias
                linhas.append(f"## {produto.categoria}")
                linhas.append("")
                categoria_atual = produto.categoria
            
            # Nome do produto
            linhas.append(produto.nome)
            
            # Descrição (se houver)
            if produto.descricao:
                linhas.append(produto.descricao)
            
            # Preço (verificar se tem promoção ativa)
            promo = promocao_por_produto.get(produto.id)
            if promo and promo.preco:
                # Produto com promoção
                preco_original = produto.preco or promo.preco_original
                desconto = promo.get_discount_percentage()
                if preco_original and desconto:
                    linhas.append(f"Preço: R$ {promo.preco:.2f} (de R$ {preco_original:.2f} - {int(desconto)}% OFF!)")
                else:
                    linhas.append(f"Preço: R$ {promo.preco:.2f} (PROMOÇÃO)")
                
                # Validade da promoção
                if promo.data_fim:
                    linhas.append(f"Promoção válida até: {promo.data_fim.strftime('%d/%m/%Y')}")
            elif produto.preco:
                linhas.append(f"Preço: R$ {produto.preco:.2f}")
            
            # URL da imagem (apenas se o produto tiver imagem)
            if produto.imagem:
                # Usar o caminho real da imagem
                linhas.append(f"Imagem do produto: https://pandia.com.br{produto.imagem.url}")
            
            linhas.append("")  # Linha em branco entre produtos
        
        extracted_text = "\n".join(linhas)
        
        # Preparar payload no mesmo formato do PDF
        rag_table_name = f"rag_{padaria.slug.replace('-', '_')}"
        
        payload = {
            "agent_id": agent.id,
            "agent_name": agent.name,
            "agent_slug": agent.slug,
            "rag_table_name": rag_table_name,
            "pdf_filename": None,
            "pdf_category": "Produtos",
            "extracted_text": extracted_text,
            "text_length": len(extracted_text),
            "padaria": padaria.name,
            "padaria_slug": padaria.slug,
            "uploaded_by": user.email if user else "",
            "action": action,
            "total_produtos": produtos.count(),
            "produtos_com_promocao": len(promocao_por_produto)
        }
        
        # Enviar para webhook N8N
        webhook_url = "https://n8n.newcouros.com.br/webhook/memoria_pandia"
        
        response = requests.post(
            webhook_url,
            json=payload,
            headers={'Content-Type': 'application/json'},
            timeout=10
        )
        
        if response.status_code == 200:
            print(f"[DEBUG] Webhook de produtos enviado com sucesso! {produtos.count()} produtos, {len(extracted_text)} caracteres.")
            return True
        else:
            print(f"[WARNING] Webhook de produtos retornou status {response.status_code}")
            return False
            
    except requests.exceptions.RequestException as e:
        print(f"[WARNING] Erro ao enviar webhook de produtos: {str(e)}")
        return False
    except Exception as e:
        print(f"[ERROR] Erro inesperado ao enviar webhook de produtos: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return False


@login_required
def organization_list(request):
    """Redireciona para a padaria do usuário ou lista se for admin."""
    padarias = get_user_padarias(request.user)
    
    # Se não for superuser e tiver apenas uma padaria, redireciona direto
    if not request.user.is_superuser and padarias.count() == 1:
        padaria = padarias.first()
        return redirect('organizations:detail', slug=padaria.slug)
    
    return render(request, "organizations/list.html", {"organizations": padarias})


@login_required
def organization_detail(request, slug):
    """
    Detalhe de uma padaria.
    
    CORREÇÃO: O problema estava aqui - o objeto precisa ser passado
    corretamente no contexto do template.
    """
    # Buscar padarias com otimização (select_related/prefetch_related se necessário)
    padarias = get_user_padarias(request.user)
    
    # Buscar a organização específica
    organization = get_object_or_404(Padaria, slug=slug)
    
    # Verificar acesso
    if not request.user.is_superuser and organization not in padarias:
        messages.error(request, "Você não tem acesso a esta padaria.")
        return redirect("organizations:list")
    
    # CORREÇÃO: Passar o objeto correto no contexto
    context = {
        'organization': organization,  # Nome correto da variável
    }
    
    # CORREÇÃO: Caminho correto do template
    return render(request, "organizations/detail.html", context)


@login_required
def organization_create(request):
    """Criar nova padaria (apenas admin)."""
    if not request.user.is_superuser:
        messages.error(request, "Apenas administradores podem criar padarias.")
        return redirect("organizations:list")
    
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        cnpj = request.POST.get("cnpj", "").strip()
        phone = request.POST.get("phone", "").strip()
        email = request.POST.get("email", "").strip()
        address = request.POST.get("address", "").strip()
        
        if not name:
            messages.error(request, "Nome é obrigatório.")
            return render(request, "organizations/form.html")
        
        try:
            # Criar padaria
            org = Padaria.objects.create(
                name=name,
                owner=request.user,
                cnpj=cnpj or None,  # Evitar strings vazias
                phone=phone or None,
                email=email or None,
                address=address or None
            )
            
            # Criar membership como dono
            PadariaUser.objects.create(
                user=request.user,
                padaria=org,
                role='dono'
            )
            
            # Log de auditoria
            AuditLog.log(
                action="create",
                entity="Padaria",
                padaria=org,
                actor=request.user,
                entity_id=org.id,
                diff={
                    "name": name,
                    "cnpj": cnpj,
                    "phone": phone,
                    "email": email,
                    "address": address
                }
            )
            
            messages.success(
                request,
                f"Padaria '{name}' criada com sucesso!\n"
                f"Usuário '{request.user.email}' vinculado como dono."
            )
            return redirect("organizations:detail", slug=org.slug)
            
        except Exception as e:
            messages.error(request, f"Erro ao criar padaria: {str(e)}")
            return render(request, "organizations/form.html")
    
    return render(request, "organizations/form.html")


@login_required
def organization_edit(request, slug):
    """Editar padaria."""
    padarias = get_user_padarias(request.user)
    organization = get_object_or_404(Padaria, slug=slug)
    
    # Verificar acesso
    if not request.user.is_superuser and organization not in padarias:
        messages.error(request, "Você não tem acesso a esta padaria.")
        return redirect("organizations:list")
    
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        cnpj = request.POST.get("cnpj", "").strip()
        phone = request.POST.get("phone", "").strip()
        email = request.POST.get("email", "").strip()
        address = request.POST.get("address", "").strip()
        
        if not name:
            messages.error(request, "Nome é obrigatório.")
            return render(request, "organizations/form.html", {"organization": organization})
        
        try:
            # Preparar diff para auditoria
            diff = {}
            
            if organization.name != name:
                diff['name'] = {'old': organization.name, 'new': name}
            if organization.cnpj != cnpj:
                diff['cnpj'] = {'old': organization.cnpj or '', 'new': cnpj}
            if organization.phone != phone:
                diff['phone'] = {'old': organization.phone or '', 'new': phone}
            if organization.email != email:
                diff['email'] = {'old': organization.email or '', 'new': email}
            if organization.address != address:
                diff['address'] = {'old': organization.address or '', 'new': address}
            
            # Atualizar campos
            organization.name = name
            organization.cnpj = cnpj or None
            organization.phone = phone or None
            organization.email = email or None
            organization.address = address or None
            organization.save()
            
            # Log de auditoria apenas se houve mudanças
            if diff:
                AuditLog.log(
                    action="update",
                    entity="Padaria",
                    padaria=organization,
                    actor=request.user,
                    entity_id=organization.id,
                    diff=diff
                )
            
            messages.success(request, "Informações atualizadas com sucesso!")
            return redirect("organizations:detail", slug=organization.slug)
            
        except Exception as e:
            messages.error(request, f"Erro ao atualizar padaria: {str(e)}")
    
    context = {'organization': organization}
    return render(request, "organizations/form.html", context)


@login_required
def organization_delete(request, slug):
    """Deletar padaria (apenas admin)."""
    if not request.user.is_superuser:
        messages.error(request, "Apenas administradores podem deletar padarias.")
        return redirect("organizations:list")
    
    organization = get_object_or_404(Padaria, slug=slug)
    
    if request.method == "POST":
        org_name = organization.name
        org_id = organization.id
        
        try:
            # Log antes de deletar
            AuditLog.log(
                action="delete",
                entity="Padaria",
                padaria=None,  # Não passar a instância pois será deletada
                actor=request.user,
                entity_id=org_id,
                diff={"name": org_name, "slug": slug}
            )
            
            organization.delete()
            messages.success(request, f"Padaria '{org_name}' deletada com sucesso!")
            return redirect("organizations:list")
            
        except Exception as e:
            messages.error(request, f"Erro ao deletar padaria: {str(e)}")
            return redirect("organizations:detail", slug=slug)
    
    context = {'organization': organization}
    return render(request, "organizations/confirm_delete.html", context)


@login_required
def apikey_list(request):
    """Lista de API keys do usuário."""
    padarias = get_user_padarias(request.user)
    
    # Admin vê todas as API Keys das padarias dele
    if request.user.is_superuser:
        api_keys = ApiKey.objects.filter(
            padaria__in=padarias
        ).select_related('padaria', 'agent')
    else:
        # Usuários normais só veem API Keys dos seus agentes
        from agents.models import Agent
        user_agents = Agent.objects.filter(padaria__in=padarias)
        api_keys = ApiKey.objects.filter(
            agent__in=user_agents
        ).select_related('padaria', 'agent')
    
    context = {'api_keys': api_keys}
    return render(request, "organizations/apikey_list.html", context)


@login_required
def apikey_create(request):
    """Criar nova API key."""
    padarias = get_user_padarias(request.user)
    
    if request.method == "POST":
        org_id = request.POST.get("organization")
        agent_id = request.POST.get("agent")
        name = request.POST.get("name", "").strip()
        
        if not org_id:
            messages.error(request, "Selecione uma organização.")
            return redirect("organizations:apikey_create")
        
        organization = get_object_or_404(Padaria, id=org_id)
        
        # Verificar acesso
        if not request.user.is_superuser and organization not in padarias:
            messages.error(request, "Você não tem acesso a esta padaria.")
            return redirect("organizations:apikeys")
        
        # Para usuários não-admin, agente é obrigatório
        if not request.user.is_superuser and not agent_id:
            messages.error(request, "Você deve selecionar um agente específico para a API Key.")
            return redirect("organizations:apikey_create")
        
        # Buscar agente se fornecido
        agent = None
        if agent_id:
            from agents.models import Agent
            agent = get_object_or_404(Agent, id=agent_id, padaria=organization)
            
            # Verificar se usuário tem acesso a este agente (não-admin)
            if not request.user.is_superuser:
                user_padarias = PadariaUser.objects.filter(
                    user=request.user
                ).values_list('padaria_id', flat=True)
                
                if agent.padaria_id not in user_padarias:
                    messages.error(request, "Você não tem acesso a este agente.")
                    return redirect("organizations:apikeys")
        
        try:
            api_key = ApiKey.objects.create(
                padaria=organization,
                agent=agent,
                name=name
            )
            
            agent_info = f" para agente '{agent.name}'" if agent else " (acesso a todos os agentes)"
            
            AuditLog.log(
                action="create",
                entity="ApiKey",
                padaria=organization,
                actor=request.user,
                entity_id=api_key.id,
                diff={
                    "name": name,
                    "agent": agent.name if agent else "Todos"
                }
            )
            
            messages.success(request, f"API Key criada{agent_info}: {api_key.key}")
            messages.warning(request, "⚠️ Copie a chave agora! Ela não será exibida novamente.")
            return redirect("organizations:apikeys")
            
        except Exception as e:
            messages.error(request, f"Erro ao criar API Key: {str(e)}")
            return redirect("organizations:apikey_create")
    
    # Buscar agentes para o formulário
    from agents.models import Agent
    agents_by_org = {}
    
    for padaria in padarias:
        agents_by_org[padaria.id] = list(
            Agent.objects.filter(
                padaria=padaria,
                is_active=True
            ).values('id', 'name')
        )
    
    context = {
        "organizations": padarias,
        "agents_by_org": agents_by_org,
        "is_admin": request.user.is_superuser
    }
    
    return render(request, "organizations/apikey_form.html", context)


@login_required
def apikey_delete(request, pk):
    """Deletar API key."""
    padarias = get_user_padarias(request.user)
    api_key = get_object_or_404(ApiKey, pk=pk)
    
    # Verificar acesso - admin pode deletar qualquer uma da padaria
    if request.user.is_superuser:
        if api_key.padaria not in padarias:
            messages.error(request, "Você não tem acesso a esta API key.")
            return redirect("organizations:apikeys")
    else:
        # Usuários normais só podem deletar API Keys dos seus agentes
        from agents.models import Agent
        user_agents = Agent.objects.filter(padaria__in=padarias)
        
        if not api_key.agent or api_key.agent not in user_agents:
            messages.error(request, "Você não tem acesso a esta API key.")
            return redirect("organizations:apikeys")
    
    if request.method == "POST":
        organization = api_key.padaria
        key_preview = api_key.key[:12]
        api_key_id = api_key.id
        
        try:
            AuditLog.log(
                action="delete",
                entity="ApiKey",
                padaria=organization,
                actor=request.user,
                entity_id=api_key_id,
                diff={"key_preview": key_preview}
            )
            
            api_key.delete()
            messages.success(request, "API Key deletada com sucesso!")
            return redirect("organizations:apikeys")
            
        except Exception as e:
            messages.error(request, f"Erro ao deletar API Key: {str(e)}")
            return redirect("organizations:apikeys")
    
    context = {'api_key': api_key}
    return render(request, "organizations/apikey_confirm_delete.html", context)


@login_required
def whatsapp_connect(request, slug):
    """Gerar QR Code para conectar WhatsApp via Evolution API."""
    padarias = get_user_padarias(request.user)
    organization = get_object_or_404(Padaria, slug=slug)
    
    # Verificar acesso
    if not request.user.is_superuser and organization not in padarias:
        messages.error(request, "Você não tem acesso a esta padaria.")
        return redirect("organizations:list")
    
    # Nome da instância baseado no slug da padaria
    instance_name = f"padaria_{organization.slug}"
    
    context = {
        "organization": organization,
        "instance_name": instance_name,
        "qr_code": None,
        "is_connected": False,
        "error": None,
        "pairing_code": None,
    }
    
    # Se for requisição AJAX para buscar QR Code
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            # Configurar Evolution API
            api_url = getattr(settings, 'EVOLUTION_API_URL', None)
            api_key = getattr(settings, 'EVOLUTION_API_KEY', None)
            
            if not api_url or not api_key:
                return JsonResponse({
                    "success": False,
                    "error": "Sistema não configurado corretamente. Entre em contato com o suporte."
                })
            
            headers = {
                "apikey": api_key,
                "Content-Type": "application/json"
            }
            
            # Log debug da URL sendo usada
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"[Evolution API] URL base: {api_url}")
            logger.info(f"[Evolution API] Tentando criar/conectar instância: {instance_name}")
            
            # Passo 1: Tentar criar a instância (se não existir)
            create_url = f"{api_url}/instance/create"
            create_payload = {
                "instanceName": instance_name,
                "qrcode": True,
                "integration": "WHATSAPP-BAILEYS"
            }
            
            evolution_hash = None  # Variável para armazenar o hash
            
            try:
                create_response = requests.post(
                    create_url,
                    headers=headers,
                    json=create_payload,
                    timeout=10
                )
                
                # Se retornar 201 (criado) ou 200 (já existe), seguir em frente
                if create_response.status_code in [200, 201]:
                    # Extrair hash da resposta
                    try:
                        create_data = create_response.json()
                        evolution_hash = (
                            create_data.get("hash") or 
                            create_data.get("token") or
                            create_data.get("instance", {}).get("token") or
                            create_data.get("instance", {}).get("hash")
                        )
                        
                        # Salvar hash no Supabase se capturado
                        if evolution_hash:
                            from integrations.supabase_client import update_agent_evolution_hash
                            # O slug da padaria é usado para encontrar o agente
                            update_agent_evolution_hash(organization.slug, evolution_hash)
                            
                            # Log para debug
                            import logging
                            logger = logging.getLogger(__name__)
                            logger.info(f"Hash Evolution capturado: {evolution_hash[:20]}... para {organization.slug}")
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(f"Não foi possível extrair hash da Evolution API: {str(e)}")
                    
                    # Registrar criação da instância
                    AuditLog.log(
                        action="whatsapp_instance_created",
                        entity="Padaria",
                        padaria=organization,
                        actor=request.user,
                        entity_id=organization.id,
                        diff={"instance": instance_name, "hash_captured": bool(evolution_hash)}
                    )
                    
                    # Configurar webhook automaticamente
                    _configure_webhook(api_url, headers, instance_name)
                    
                elif create_response.status_code in [403, 409]:
                    # Instância já existe, está OK - configurar webhook mesmo assim
                    _configure_webhook(api_url, headers, instance_name)
                else:
                    # Outro erro ao criar
                    error_msg = _extract_error_message(create_response)
                    return JsonResponse({
                        "success": False,
                        "error": f"Erro ao criar instância: {error_msg}"
                    })
                    
            except requests.exceptions.RequestException:
                # Se falhar ao criar, tentar conectar (pode já existir)
                pass
            
            # Passo 2: Buscar QR Code da instância
            connect_url = f"{api_url}/instance/connect/{instance_name}"
            response = requests.get(connect_url, headers=headers, timeout=10)
            
            # Tratar resposta
            if response.status_code == 200:
                data = response.json()
                
                # Verificar se já está conectado
                instance_data = data.get("instance", {})
                state = data.get("state") or instance_data.get("state")
                status = data.get("status") or instance_data.get("status")
                
                if state == "open" or status == "connected":
                    return JsonResponse({
                        "success": True,
                        "is_connected": True,
                        "message": "WhatsApp já está conectado!"
                    })
                
                # Extrair QR Code e pairing code
                qr_code = _extract_qr_code(data)
                pairing_code = _extract_pairing_code(data)
                
                if qr_code:
                    # Adicionar prefixo data:image se necessário
                    if not qr_code.startswith("data:image"):
                        qr_code = f"data:image/png;base64,{qr_code}"
                    
                    # Registrar auditoria
                    AuditLog.log(
                        action="whatsapp_qr_generated",
                        entity="Padaria",
                        padaria=organization,
                        actor=request.user,
                        entity_id=organization.id,
                        diff={"instance": instance_name}
                    )
                    
                    return JsonResponse({
                        "success": True,
                        "qr_code": qr_code,
                        "pairing_code": pairing_code,
                        "is_connected": False
                    })
                else:
                    # Aguardando inicialização
                    return JsonResponse({
                        "success": False,
                        "error": "O WhatsApp ainda está inicializando. "
                                "Por favor, aguarde alguns segundos e tente novamente."
                    })
                    
            elif response.status_code == 404:
                return JsonResponse({
                    "success": False,
                    "error": "Instância WhatsApp não encontrada. "
                            "Tente recarregar a página e gerar novamente."
                })
            else:
                error_msg = _extract_error_message(response)
                return JsonResponse({
                    "success": False,
                    "error": error_msg
                })
                
        except requests.exceptions.Timeout:
            return JsonResponse({
                "success": False,
                "error": "A conexão demorou muito para responder. "
                        "Verifique sua internet e tente novamente."
            })
        except requests.exceptions.ConnectionError as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"[Evolution API] ConnectionError: {str(e)}")
            logger.error(f"[Evolution API] URL usada: {api_url}")
            
            return JsonResponse({
                "success": False,
                "error": f"Não foi possível conectar ao serviço WhatsApp. "
                        f"Verifique se a Evolution API está acessível em: {api_url}"
            })
        except Exception as e:
            # Log do erro real para debug
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro no WhatsApp connect: {str(e)}", exc_info=True)
            
            return JsonResponse({
                "success": False,
                "error": "Ocorreu um erro inesperado. Por favor, tente novamente."
            })
    
    # Renderizar página
    return render(request, "organizations/whatsapp_connect.html", context)


# Funções auxiliares
def _extract_error_message(response):
    """Extrai mensagem de erro de forma amigável."""
    try:
        error_data = response.json()
        
        # Tentar extrair de response.message primeiro
        if isinstance(error_data.get("response"), dict):
            error_msg = error_data["response"].get("message", "")
        else:
            error_msg = error_data.get("message", "")
        
        # Se message é uma lista, extrair primeiro item
        if isinstance(error_msg, list) and len(error_msg) > 0:
            error_msg = error_msg[0]
        
        if not error_msg:
            error_msg = response.text or "Erro desconhecido"
            
        return error_msg
    except Exception:
        return "Ocorreu um erro inesperado. Tente novamente."


def _extract_qr_code(data):
    """Extrai QR Code de diferentes estruturas de resposta."""
    qr_code = None
    
    # Verificar diferentes estruturas
    if "base64" in data:
        qr_code = data["base64"]
    elif "qrcode" in data:
        qr_code = data["qrcode"]
        if isinstance(qr_code, dict):
            qr_code = qr_code.get("base64") or qr_code.get("code")
    elif "code" in data:
        qr_code = data["code"]
    
    return qr_code


def _extract_pairing_code(data):
    """Extrai pairing code da resposta."""
    pairing_code = None
    
    if "pairingCode" in data:
        pairing_code = data["pairingCode"]
    elif "code" in data and not _extract_qr_code(data):
        pairing_code = data["code"]
    
    return pairing_code


@login_required
def whatsapp_pairing_code(request, slug):
    """Gerar Pairing Code para conectar WhatsApp via número de telefone."""
    from django.views.decorators.http import require_POST
    
    padarias = get_user_padarias(request.user)
    organization = get_object_or_404(Padaria, slug=slug)
    
    # Verificar acesso
    if not request.user.is_superuser and organization not in padarias:
        return JsonResponse({
            "success": False,
            "error": "Você não tem acesso a esta padaria."
        })
    
    # Apenas POST
    if request.method != "POST":
        return JsonResponse({
            "success": False,
            "error": "Método não permitido."
        })
    
    # Obter número de telefone do body
    import json
    try:
        data = json.loads(request.body)
        phone_number = data.get("phone", "").strip()
    except:
        phone_number = request.POST.get("phone", "").strip()
    
    # Validar número
    phone_number = ''.join(filter(str.isdigit, phone_number))
    
    if not phone_number:
        return JsonResponse({
            "success": False,
            "error": "Número de telefone é obrigatório."
        })
    
    if len(phone_number) < 10:
        return JsonResponse({
            "success": False,
            "error": "Número de telefone inválido. Inclua o código do país (ex: 5511999998888)."
        })
    
    # Nome da instância
    instance_name = f"padaria_{organization.slug}"
    
    try:
        # Configurar Evolution API
        api_url = getattr(settings, 'EVOLUTION_API_URL', None)
        api_key = getattr(settings, 'EVOLUTION_API_KEY', None)
        
        if not api_url or not api_key:
            return JsonResponse({
                "success": False,
                "error": "Sistema não configurado corretamente. Entre em contato com o suporte."
            })
        
        headers = {
            "apikey": api_key,
            "Content-Type": "application/json"
        }
        
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"[Evolution API] Gerando pairing code para {instance_name} com número {phone_number[:4]}***")
        
        # Chamar endpoint para gerar pairing code
        # GET /instance/connect/{instance}?number={phone}
        connect_url = f"{api_url}/instance/connect/{instance_name}"
        response = requests.get(
            connect_url, 
            headers=headers, 
            params={"number": phone_number},
            timeout=15
        )
        
        logger.info(f"[Evolution API] Resposta status: {response.status_code}")
        
        if response.status_code == 200:
            data = response.json()
            
            # Log para debug
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"[Evolution API] Resposta pairing code: {data}")
            
            # Verificar se já está conectado
            instance_data = data.get("instance", {})
            state = data.get("state") or instance_data.get("state")
            status = data.get("status") or instance_data.get("status")
            
            if state == "open" or status == "connected":
                return JsonResponse({
                    "success": True,
                    "is_connected": True,
                    "message": "WhatsApp já está conectado!"
                })
            
            # Extrair pairing code diretamente
            pairing_code = data.get("pairingCode")
            code_field = data.get("code")
            
            # Log para debug
            logger.info(f"[Evolution API] pairingCode: {pairing_code}")
            logger.info(f"[Evolution API] code: {code_field[:50] if code_field else None}...")
            
            if pairing_code:
                # Se o código tem 8 caracteres sem hífen, formatar
                clean_code = str(pairing_code).replace("-", "").replace(" ", "")
                if len(clean_code) == 8:
                    pairing_code = f"{clean_code[:4]}-{clean_code[4:]}"
                
                # Registrar auditoria
                AuditLog.log(
                    action="whatsapp_pairing_code_generated",
                    entity="Padaria",
                    padaria=organization,
                    actor=request.user,
                    entity_id=organization.id,
                    diff={"instance": instance_name, "phone_prefix": phone_number[:4]}
                )
                
                return JsonResponse({
                    "success": True,
                    "pairing_code": pairing_code,
                    "is_connected": False
                })
            else:
                # Mostrar todos os valores para debug
                debug_info = {
                    "pairingCode": data.get("pairingCode"),
                    "code_preview": str(data.get("code", ""))[:30] if data.get("code") else None,
                    "count": data.get("count"),
                    "keys": list(data.keys())
                }
                logger.warning(f"[Evolution API] Debug: {debug_info}")
                return JsonResponse({
                    "success": False,
                    "error": f"A API não retornou o código de pareamento. O recurso de pairing code pode não estar disponível nesta versão da Evolution API."
                })
        
        elif response.status_code == 404:
            return JsonResponse({
                "success": False,
                "error": "Instância WhatsApp não encontrada. Gere o QR Code primeiro para criar a instância."
            })
        else:
            error_msg = _extract_error_message(response)
            return JsonResponse({
                "success": False,
                "error": error_msg
            })
            
    except requests.exceptions.Timeout:
        return JsonResponse({
            "success": False,
            "error": "A conexão demorou muito para responder. Tente novamente."
        })
    except requests.exceptions.ConnectionError:
        return JsonResponse({
            "success": False,
            "error": "Não foi possível conectar ao serviço WhatsApp."
        })
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Erro no WhatsApp pairing code: {str(e)}", exc_info=True)
        
        return JsonResponse({
            "success": False,
            "error": "Ocorreu um erro inesperado. Por favor, tente novamente."
        })


# =====================================================
# PROMOÇÕES E AVISOS
# =====================================================

@login_required
def promocao_list(request):
    """Lista de promoções da padaria."""
    padarias = get_user_padarias(request.user)
    
    if not padarias.exists():
        messages.warning(request, "Você não tem acesso a nenhuma padaria.")
        return redirect("organizations:list")
    
    # Para usuários normais, pegar a primeira padaria
    # Para admin, mostrar todas as promoções
    if request.user.is_superuser:
        promocoes = Promocao.objects.filter(
            padaria__in=padarias
        ).select_related('padaria').order_by('-created_at')
    else:
        padaria = padarias.first()
        promocoes = Promocao.objects.filter(
            padaria=padaria
        ).order_by('-created_at')
    
    context = {
        'promocoes': promocoes,
        'padarias': padarias,
    }
    return render(request, "organizations/promocao_list.html", context)


@login_required
def promocao_create(request):
    """Criar nova promoção."""
    padarias = get_user_padarias(request.user)
    
    if not padarias.exists():
        messages.error(request, "Você não tem acesso a nenhuma padaria.")
        return redirect("organizations:list")
    
    # Pegar a primeira padaria do usuário (ou a única)
    padaria = padarias.first()
    
    # Buscar produtos da padaria para o dropdown
    produtos = Produto.objects.filter(padaria=padaria, ativo=True).order_by('nome')
    
    if request.method == "POST":
        titulo = request.POST.get("titulo", "").strip()
        descricao = request.POST.get("descricao", "").strip()
        preco = request.POST.get("preco", "").strip()
        preco_original = request.POST.get("preco_original", "").strip()
        data_inicio = request.POST.get("data_inicio", "").strip()
        data_fim = request.POST.get("data_fim", "").strip()
        is_active = request.POST.get("is_active") == "on"
        imagem = request.FILES.get("imagem")
        produto_id = request.POST.get("produto", "").strip()
        
        if not titulo:
            messages.error(request, "O título é obrigatório.")
            return render(request, "organizations/promocao_form.html", {
                "padarias": padarias,
                "produtos": produtos,
            })
        
        try:
            # Buscar produto vinculado se informado
            produto_vinculado = None
            if produto_id:
                try:
                    produto_vinculado = Produto.objects.get(pk=produto_id, padaria=padaria)
                except Produto.DoesNotExist:
                    pass
            
            promocao = Promocao.objects.create(
                padaria=padaria,
                produto=produto_vinculado,
                titulo=titulo,
                descricao=descricao,
                preco=float(preco.replace(",", ".")) if preco else None,
                preco_original=float(preco_original.replace(",", ".")) if preco_original else None,
                data_inicio=data_inicio if data_inicio else None,
                data_fim=data_fim if data_fim else None,
                is_active=is_active,
                imagem=imagem
            )
            
            AuditLog.log(
                action="create",
                entity="Promocao",
                padaria=padaria,
                actor=request.user,
                entity_id=promocao.id,
                diff={"titulo": titulo}
            )
            
            # Se promoção está vinculada a produto, atualizar RAG
            if produto_vinculado:
                send_products_webhook(padaria, request.user, action="promotion_created")
            
            messages.success(request, f"Promoção '{titulo}' criada com sucesso!")
            return redirect("organizations:promocao_list")
            
        except Exception as e:
            messages.error(request, f"Erro ao criar promoção: {str(e)}")
            return render(request, "organizations/promocao_form.html", {
                "padarias": padarias,
                "produtos": produtos,
            })
    
    context = {
        "padarias": padarias,
        "produtos": produtos,
    }
    return render(request, "organizations/promocao_form.html", context)


@login_required
def promocao_edit(request, pk):
    """Editar promoção existente."""
    padarias = get_user_padarias(request.user)
    promocao = get_object_or_404(Promocao, pk=pk)
    
    # Verificar acesso
    if not request.user.is_superuser and promocao.padaria not in padarias:
        messages.error(request, "Você não tem acesso a esta promoção.")
        return redirect("organizations:promocao_list")
    
    # Buscar produtos da padaria para o dropdown
    produtos = Produto.objects.filter(padaria=promocao.padaria, ativo=True).order_by('nome')
    
    if request.method == "POST":
        titulo = request.POST.get("titulo", "").strip()
        descricao = request.POST.get("descricao", "").strip()
        preco = request.POST.get("preco", "").strip()
        preco_original = request.POST.get("preco_original", "").strip()
        data_inicio = request.POST.get("data_inicio", "").strip()
        data_fim = request.POST.get("data_fim", "").strip()
        is_active = request.POST.get("is_active") == "on"
        imagem = request.FILES.get("imagem")
        remover_imagem = request.POST.get("remover_imagem") == "on"
        produto_id = request.POST.get("produto", "").strip()
        
        if not titulo:
            messages.error(request, "O título é obrigatório.")
            return render(request, "organizations/promocao_form.html", {
                "promocao": promocao,
                "padarias": padarias,
                "produtos": produtos,
            })
        
        try:
            # Preparar diff para auditoria
            diff = {}
            if promocao.titulo != titulo:
                diff['titulo'] = {'old': promocao.titulo, 'new': titulo}
            
            # Atualizar campos
            promocao.titulo = titulo
            promocao.descricao = descricao
            promocao.preco = float(preco.replace(",", ".")) if preco else None
            promocao.preco_original = float(preco_original.replace(",", ".")) if preco_original else None
            promocao.data_inicio = data_inicio if data_inicio else None
            promocao.data_fim = data_fim if data_fim else None
            promocao.is_active = is_active
            
            # Atualizar produto vinculado
            if produto_id:
                try:
                    promocao.produto = Produto.objects.get(pk=produto_id, padaria=promocao.padaria)
                except Produto.DoesNotExist:
                    promocao.produto = None
            else:
                promocao.produto = None
            
            if remover_imagem:
                promocao.imagem = None
            elif imagem:
                promocao.imagem = imagem
            
            promocao.save()
            
            if diff:
                AuditLog.log(
                    action="update",
                    entity="Promocao",
                    padaria=promocao.padaria,
                    actor=request.user,
                    entity_id=promocao.id,
                    diff=diff
                )
            
            # Se promoção está vinculada a produto, atualizar RAG
            if promocao.produto:
                send_products_webhook(promocao.padaria, request.user, action="promotion_updated")
            
            messages.success(request, "Promoção atualizada com sucesso!")
            return redirect("organizations:promocao_list")
            
        except Exception as e:
            messages.error(request, f"Erro ao atualizar promoção: {str(e)}")
    
    context = {
        "promocao": promocao,
        "padarias": padarias,
        "produtos": produtos,
    }
    return render(request, "organizations/promocao_form.html", context)


@login_required
def promocao_delete(request, pk):
    """Deletar promoção."""
    padarias = get_user_padarias(request.user)
    promocao = get_object_or_404(Promocao, pk=pk)
    
    # Verificar acesso
    if not request.user.is_superuser and promocao.padaria not in padarias:
        messages.error(request, "Você não tem acesso a esta promoção.")
        return redirect("organizations:promocao_list")
    
    if request.method == "POST":
        titulo = promocao.titulo
        promocao_id = promocao.id
        padaria = promocao.padaria
        tinha_produto = promocao.produto is not None  # Guardar antes de deletar
        
        try:
            AuditLog.log(
                action="delete",
                entity="Promocao",
                padaria=padaria,
                actor=request.user,
                entity_id=promocao_id,
                diff={"titulo": titulo}
            )
            
            promocao.delete()
            
            # Se promoção estava vinculada a produto, atualizar RAG
            if tinha_produto:
                send_products_webhook(padaria, request.user, action="promotion_deleted")
            
            messages.success(request, f"Promoção '{titulo}' excluída com sucesso!")
            return redirect("organizations:promocao_list")
            
        except Exception as e:
            messages.error(request, f"Erro ao excluir promoção: {str(e)}")
            return redirect("organizations:promocao_list")
    
    context = {'promocao': promocao}
    return render(request, "organizations/promocao_confirm_delete.html", context)


# =====================================================
# PRODUTOS
# =====================================================

@login_required
def produto_list(request):
    """Lista de produtos da padaria."""
    padarias = get_user_padarias(request.user)
    
    if not padarias.exists():
        messages.warning(request, "Você não tem acesso a nenhuma padaria.")
        return redirect("organizations:list")
    
    # Para usuários normais, pegar a primeira padaria
    # Para admin, mostrar todos os produtos
    if request.user.is_superuser:
        produtos = Produto.objects.filter(
            padaria__in=padarias
        ).select_related('padaria').order_by('categoria', 'nome')
        # Buscar promoções ativas de todas as padarias
        promocoes_ativas = Promocao.objects.filter(
            padaria__in=padarias,
            produto__isnull=False,
            is_active=True
        ).select_related('produto')
    else:
        padaria = padarias.first()
        produtos = Produto.objects.filter(
            padaria=padaria
        ).order_by('categoria', 'nome')
        # Buscar promoções ativas da padaria
        promocoes_ativas = Promocao.objects.filter(
            padaria=padaria,
            produto__isnull=False,
            is_active=True
        ).select_related('produto')
    
    # Criar dicionário de produto_id -> promocao para lookup rápido
    promocao_por_produto = {}
    for promo in promocoes_ativas:
        if promo.is_valid() and promo.produto_id:
            # Se já tem uma promoção para este produto, manter a com maior desconto
            if promo.produto_id not in promocao_por_produto:
                promocao_por_produto[promo.produto_id] = promo
    
    # Anexar promoção ativa a cada produto
    produtos_list = list(produtos)
    for produto in produtos_list:
        produto.promocao_ativa = promocao_por_produto.get(produto.id)
    
    context = {
        'produtos': produtos_list,
        'padarias': padarias,
    }
    return render(request, "organizations/produto_list.html", context)


@login_required
def produto_create(request):
    """Criar novo produto."""
    padarias = get_user_padarias(request.user)
    
    if not padarias.exists():
        messages.error(request, "Você não tem acesso a nenhuma padaria.")
        return redirect("organizations:list")
    
    # Pegar a primeira padaria do usuário (ou a única)
    padaria = padarias.first()
    
    if request.method == "POST":
        nome = request.POST.get("nome", "").strip()
        descricao = request.POST.get("descricao", "").strip()
        preco = request.POST.get("preco", "").strip()
        categoria = request.POST.get("categoria", "").strip()
        ativo = request.POST.get("ativo") == "on"
        imagem = request.FILES.get("imagem")
        
        if not nome:
            messages.error(request, "O nome é obrigatório.")
            return render(request, "organizations/produto_form.html", {
                "padarias": padarias,
            })
        
        try:
            produto = Produto.objects.create(
                padaria=padaria,
                nome=nome,
                descricao=descricao,
                preco=float(preco.replace(",", ".")) if preco else None,
                categoria=categoria,
                ativo=ativo,
                imagem=imagem
            )
            
            AuditLog.log(
                action="create",
                entity="Produto",
                padaria=padaria,
                actor=request.user,
                entity_id=produto.id,
                diff={"nome": nome}
            )
            
            # Enviar webhook para atualizar RAG do N8N
            send_products_webhook(padaria, request.user, action="product_created")
            
            messages.success(request, f"Produto '{nome}' criado com sucesso!")
            return redirect("organizations:produto_list")
            
        except Exception as e:
            messages.error(request, f"Erro ao criar produto: {str(e)}")
            return render(request, "organizations/produto_form.html", {
                "padarias": padarias,
            })
    
    context = {
        "padarias": padarias,
    }
    return render(request, "organizations/produto_form.html", context)


@login_required
def produto_edit(request, pk):
    """Editar produto existente."""
    padarias = get_user_padarias(request.user)
    produto = get_object_or_404(Produto, pk=pk)
    
    # Verificar acesso
    if not request.user.is_superuser and produto.padaria not in padarias:
        messages.error(request, "Você não tem acesso a este produto.")
        return redirect("organizations:produto_list")
    
    if request.method == "POST":
        nome = request.POST.get("nome", "").strip()
        descricao = request.POST.get("descricao", "").strip()
        preco = request.POST.get("preco", "").strip()
        categoria = request.POST.get("categoria", "").strip()
        ativo = request.POST.get("ativo") == "on"
        imagem = request.FILES.get("imagem")
        remover_imagem = request.POST.get("remover_imagem") == "on"
        
        if not nome:
            messages.error(request, "O nome é obrigatório.")
            return render(request, "organizations/produto_form.html", {
                "produto": produto,
                "padarias": padarias,
            })
        
        try:
            # Preparar diff para auditoria
            diff = {}
            if produto.nome != nome:
                diff['nome'] = {'old': produto.nome, 'new': nome}
            
            # Atualizar campos
            produto.nome = nome
            produto.descricao = descricao
            produto.preco = float(preco.replace(",", ".")) if preco else None
            produto.categoria = categoria
            produto.ativo = ativo
            
            # Gerenciar imagem
            if remover_imagem and produto.imagem:
                produto.imagem.delete(save=False)
                produto.imagem = None
            elif imagem:
                if produto.imagem:
                    produto.imagem.delete(save=False)
                produto.imagem = imagem
            
            produto.save()
            
            if diff:
                AuditLog.log(
                    action="update",
                    entity="Produto",
                    padaria=produto.padaria,
                    actor=request.user,
                    entity_id=produto.id,
                    diff=diff
                )
            
            # Enviar webhook para atualizar RAG do N8N
            send_products_webhook(produto.padaria, request.user, action="product_updated")
            
            messages.success(request, f"Produto '{nome}' atualizado com sucesso!")
            return redirect("organizations:produto_list")
            
        except Exception as e:
            messages.error(request, f"Erro ao atualizar produto: {str(e)}")
    
    context = {
        "produto": produto,
        "padarias": padarias,
    }
    return render(request, "organizations/produto_form.html", context)


@login_required
def produto_delete(request, pk):
    """Deletar produto."""
    padarias = get_user_padarias(request.user)
    produto = get_object_or_404(Produto, pk=pk)
    
    # Verificar acesso
    if not request.user.is_superuser and produto.padaria not in padarias:
        messages.error(request, "Você não tem acesso a este produto.")
        return redirect("organizations:produto_list")
    
    if request.method == "POST":
        nome = produto.nome
        produto_id = produto.id
        padaria = produto.padaria
        
        try:
            AuditLog.log(
                action="delete",
                entity="Produto",
                padaria=padaria,
                actor=request.user,
                entity_id=produto_id,
                diff={"nome": nome}
            )
            
            produto.delete()
            
            # Enviar webhook para atualizar RAG do N8N (produto removido)
            send_products_webhook(padaria, request.user, action="product_deleted")
            
            messages.success(request, f"Produto '{nome}' excluído com sucesso!")
            return redirect("organizations:produto_list")
            
        except Exception as e:
            messages.error(request, f"Erro ao excluir produto: {str(e)}")
            return redirect("organizations:produto_list")
    
    context = {'produto': produto}
    return render(request, "organizations/produto_confirm_delete.html", context)


@login_required
def produto_import(request):
    """Importar produtos de PDF."""
    from agents.utils import extract_text_from_pdf, extract_products_from_text
    
    padarias = get_user_padarias(request.user)
    
    if not padarias.exists():
        messages.error(request, "Você não tem acesso a nenhuma padaria.")
        return redirect("organizations:list")
    
    # Pegar a primeira padaria do usuário
    padaria = padarias.first()
    
    if request.method == "POST":
        pdf_file = request.FILES.get("pdf_file")
        
        print(f"[DEBUG] produto_import POST recebido")
        print(f"[DEBUG] pdf_file: {pdf_file}")
        
        if not pdf_file:
            messages.error(request, "Por favor, selecione um arquivo PDF.")
            return render(request, "organizations/produto_import.html", {"padarias": padarias})
        
        # Verificar se é PDF
        if not pdf_file.name.lower().endswith('.pdf'):
            messages.error(request, "O arquivo deve ser um PDF.")
            return render(request, "organizations/produto_import.html", {"padarias": padarias})
        
        try:
            print(f"[DEBUG] Extraindo texto do PDF: {pdf_file.name}")
            
            # Extrair texto do PDF
            extracted_text = extract_text_from_pdf(pdf_file)
            
            print(f"[DEBUG] Texto extraído: {len(extracted_text)} caracteres")
            print(f"[DEBUG] Preview: {extracted_text[:500]}")
            
            if not extracted_text.strip():
                messages.warning(request, "Não foi possível extrair texto do PDF. Verifique se o PDF contém texto legível.")
                return render(request, "organizations/produto_import.html", {"padarias": padarias})
            
            print(f"[DEBUG] Chamando extract_products_from_text para padaria: {padaria.name}")
            
            # Extrair produtos do texto
            produtos_criados = extract_products_from_text(extracted_text, padaria)
            
            print(f"[DEBUG] Produtos criados: {len(produtos_criados) if produtos_criados else 0}")
            
            if produtos_criados:
                AuditLog.log(
                    action="import_pdf",
                    entity="Produto",
                    padaria=padaria,
                    actor=request.user,
                    entity_id=None,
                    diff={
                        "pdf_filename": pdf_file.name,
                        "produtos_importados": len(produtos_criados)
                    }
                )
                
                # Enviar webhook para atualizar RAG do N8N
                send_products_webhook(padaria, request.user, action="products_imported_pdf")
                
                messages.success(request, f"✅ {len(produtos_criados)} produtos foram importados do PDF!")
            else:
                messages.warning(request, 
                    "Nenhum produto foi encontrado no PDF. "
                    "Certifique-se de que o PDF contém produtos no formato: 'Nome - R$ 00,00'"
                )
            
            return redirect("organizations:produto_list")
            
        except Exception as e:
            import traceback
            print(f"[ERROR] Erro ao processar PDF: {str(e)}")
            print(traceback.format_exc())
            messages.error(request, f"Erro ao processar PDF: {str(e)}")
            return render(request, "organizations/produto_import.html", {"padarias": padarias})
    
    context = {
        "padarias": padarias,
    }
    return render(request, "organizations/produto_import.html", context)


@login_required
def produto_import_excel(request):
    """Importar produtos de planilha Excel."""
    from openpyxl import load_workbook
    
    padarias = get_user_padarias(request.user)
    
    if not padarias.exists():
        messages.error(request, "Você não tem acesso a nenhuma padaria.")
        return redirect("organizations:list")
    
    # Pegar a primeira padaria do usuário
    padaria = padarias.first()
    
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        
        if not excel_file:
            messages.error(request, "Por favor, selecione um arquivo Excel.")
            return render(request, "organizations/produto_import_excel.html", {"padarias": padarias})
        
        # Verificar se é Excel
        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, "O arquivo deve ser uma planilha Excel (.xlsx ou .xls).")
            return render(request, "organizations/produto_import_excel.html", {"padarias": padarias})
        
        try:
            # Carregar a planilha
            wb = load_workbook(excel_file, read_only=True)
            ws = wb.active
            
            # Ler cabeçalhos da primeira linha
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower())
                else:
                    headers.append("")
            
            # Mapear índices das colunas
            nome_idx = None
            descricao_idx = None
            preco_idx = None
            categoria_idx = None
            
            for i, h in enumerate(headers):
                h_lower = h.lower()
                # Usar 'in' para matching flexível de headers
                if any(keyword in h_lower for keyword in ['nome', 'name', 'produto', 'product']):
                    nome_idx = i
                elif any(keyword in h_lower for keyword in ['descrição', 'descricao', 'description', 'desc']):
                    descricao_idx = i
                elif any(keyword in h_lower for keyword in ['preço', 'preco', 'price', 'valor', 'value', 'r$']):
                    preco_idx = i
                elif any(keyword in h_lower for keyword in ['categoria', 'category', 'tipo', 'type']):
                    categoria_idx = i
            
            if nome_idx is None:
                messages.error(request, "Coluna 'Nome' não encontrada na planilha. Verifique o cabeçalho.")
                return render(request, "organizations/produto_import_excel.html", {"padarias": padarias})
            
            # Processar linhas (a partir da segunda)
            produtos_importados = 0
            produtos_atualizados = 0
            erros = 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Pegar valores das colunas
                    nome = str(row[nome_idx]).strip() if row[nome_idx] else None
                    
                    if not nome or nome == "None":
                        continue
                    
                    descricao = ""
                    if descricao_idx is not None and len(row) > descricao_idx and row[descricao_idx]:
                        descricao = str(row[descricao_idx]).strip()
                    
                    preco = None
                    if preco_idx is not None and len(row) > preco_idx and row[preco_idx]:
                        preco_val = row[preco_idx]
                        if isinstance(preco_val, (int, float)):
                            preco = float(preco_val)
                        else:
                            # Tentar converter string (suporta R$ 1,20 ou 1.20)
                            preco_str = str(preco_val).replace("R$", "").replace(" ", "").replace(",", ".").strip()
                            try:
                                preco = float(preco_str)
                            except ValueError:
                                preco = None
                    
                    categoria = ""
                    if categoria_idx is not None and len(row) > categoria_idx and row[categoria_idx]:
                        categoria = str(row[categoria_idx]).strip()
                    
                    # Criar ou atualizar produto
                    defaults = {
                        'descricao': descricao,
                        'preco': preco,
                        'ativo': True
                    }
                    if categoria:
                        defaults['categoria'] = categoria
                    
                    produto, created = Produto.objects.update_or_create(
                        padaria=padaria,
                        nome=nome,
                        defaults=defaults
                    )
                    
                    if created:
                        produtos_importados += 1
                    else:
                        produtos_atualizados += 1
                        
                except Exception as e:
                    print(f"[ERROR] Erro na linha {row_num}: {str(e)}")
                    erros += 1
            
            wb.close()
            
            # Registrar auditoria
            if produtos_importados > 0 or produtos_atualizados > 0:
                AuditLog.log(
                    action="import_excel",
                    entity="Produto",
                    padaria=padaria,
                    actor=request.user,
                    entity_id=None,
                    diff={
                        "excel_filename": excel_file.name,
                        "produtos_importados": produtos_importados,
                        "produtos_atualizados": produtos_atualizados,
                        "erros": erros
                    }
                )
            
            # Mensagem de resultado
            if produtos_importados > 0 or produtos_atualizados > 0:
                # Enviar webhook para atualizar RAG do N8N
                send_products_webhook(padaria, request.user, action="products_imported_excel")
                
                msg = f"✅ Importação concluída! "
                if produtos_importados > 0:
                    msg += f"{produtos_importados} produtos novos"
                if produtos_atualizados > 0:
                    if produtos_importados > 0:
                        msg += f", {produtos_atualizados} atualizados"
                    else:
                        msg += f"{produtos_atualizados} produtos atualizados"
                if erros > 0:
                    msg += f" ({erros} erros)"
                messages.success(request, msg)
            else:
                messages.warning(request, "Nenhum produto foi encontrado na planilha.")
            
            return redirect("organizations:produto_list")
            
        except Exception as e:
            import traceback
            print(f"[ERROR] Erro ao processar Excel: {str(e)}")
            print(traceback.format_exc())
            messages.error(request, f"Erro ao processar planilha: {str(e)}")
            return render(request, "organizations/produto_import_excel.html", {"padarias": padarias})
    
    context = {
        "padarias": padarias,
    }
    return render(request, "organizations/produto_import_excel.html", context)


# =====================================================
# CLIENTES
# =====================================================

@login_required
def cliente_list(request):
    """Lista de clientes da padaria."""
    padarias = get_user_padarias(request.user)
    
    if not padarias.exists():
        messages.warning(request, "Você não tem acesso a nenhuma padaria.")
        return redirect("organizations:list")
    
    if request.user.is_superuser:
        clientes = Cliente.objects.filter(
            padaria__in=padarias
        ).select_related('padaria').order_by('nome')
    else:
        padaria = padarias.first()
        clientes = Cliente.objects.filter(
            padaria=padaria
        ).order_by('nome')
    
    context = {
        'clientes': clientes,
        'padarias': padarias,
        'total_clientes': clientes.count(),
        'clientes_ativos': clientes.filter(is_active=True, aceita_promocoes=True).count(),
    }
    return render(request, "organizations/cliente_list.html", context)


@login_required
def cliente_create(request):
    """Criar novo cliente."""
    padarias = get_user_padarias(request.user)
    
    if not padarias.exists():
        messages.error(request, "Você não tem acesso a nenhuma padaria.")
        return redirect("organizations:list")
    
    padaria = padarias.first()
    
    if request.method == "POST":
        nome = request.POST.get("nome", "").strip()
        telefone = request.POST.get("telefone", "").strip()
        email = request.POST.get("email", "").strip()
        observacoes = request.POST.get("observacoes", "").strip()
        aceita_promocoes = request.POST.get("aceita_promocoes") == "on"
        
        if not nome:
            messages.error(request, "O nome é obrigatório.")
            return render(request, "organizations/cliente_form.html", {"padarias": padarias})
        
        if not telefone:
            messages.error(request, "O telefone/WhatsApp é obrigatório.")
            return render(request, "organizations/cliente_form.html", {"padarias": padarias})
        
        try:
            cliente = Cliente.objects.create(
                padaria=padaria,
                nome=nome,
                telefone=telefone,
                email=email,
                observacoes=observacoes,
                aceita_promocoes=aceita_promocoes
            )
            
            AuditLog.log(
                action="create",
                entity="Cliente",
                padaria=padaria,
                actor=request.user,
                entity_id=cliente.id,
                diff={"nome": nome, "telefone": telefone}
            )
            
            messages.success(request, f"Cliente '{nome}' cadastrado com sucesso!")
            return redirect("organizations:cliente_list")
            
        except Exception as e:
            if "unique" in str(e).lower():
                messages.error(request, "Já existe um cliente com este telefone cadastrado.")
            else:
                messages.error(request, f"Erro ao cadastrar cliente: {str(e)}")
            return render(request, "organizations/cliente_form.html", {"padarias": padarias})
    
    return render(request, "organizations/cliente_form.html", {"padarias": padarias})


@login_required
def cliente_edit(request, pk):
    """Editar cliente existente."""
    padarias = get_user_padarias(request.user)
    cliente = get_object_or_404(Cliente, pk=pk)
    
    if not request.user.is_superuser and cliente.padaria not in padarias:
        messages.error(request, "Você não tem acesso a este cliente.")
        return redirect("organizations:cliente_list")
    
    if request.method == "POST":
        nome = request.POST.get("nome", "").strip()
        telefone = request.POST.get("telefone", "").strip()
        email = request.POST.get("email", "").strip()
        observacoes = request.POST.get("observacoes", "").strip()
        aceita_promocoes = request.POST.get("aceita_promocoes") == "on"
        is_active = request.POST.get("is_active") == "on"
        
        if not nome or not telefone:
            messages.error(request, "Nome e telefone são obrigatórios.")
            return render(request, "organizations/cliente_form.html", {
                "cliente": cliente,
                "padarias": padarias,
            })
        
        try:
            cliente.nome = nome
            cliente.telefone = telefone
            cliente.email = email
            cliente.observacoes = observacoes
            cliente.aceita_promocoes = aceita_promocoes
            cliente.is_active = is_active
            cliente.save()
            
            AuditLog.log(
                action="update",
                entity="Cliente",
                padaria=cliente.padaria,
                actor=request.user,
                entity_id=cliente.id,
                diff={"nome": nome}
            )
            
            messages.success(request, "Cliente atualizado com sucesso!")
            return redirect("organizations:cliente_list")
            
        except Exception as e:
            messages.error(request, f"Erro ao atualizar cliente: {str(e)}")
    
    return render(request, "organizations/cliente_form.html", {
        "cliente": cliente,
        "padarias": padarias,
    })


@login_required
def cliente_delete(request, pk):
    """Deletar cliente."""
    padarias = get_user_padarias(request.user)
    cliente = get_object_or_404(Cliente, pk=pk)
    
    if not request.user.is_superuser and cliente.padaria not in padarias:
        messages.error(request, "Você não tem acesso a este cliente.")
        return redirect("organizations:cliente_list")
    
    if request.method == "POST":
        nome = cliente.nome
        
        try:
            AuditLog.log(
                action="delete",
                entity="Cliente",
                padaria=cliente.padaria,
                actor=request.user,
                entity_id=cliente.id,
                diff={"nome": nome}
            )
            
            cliente.delete()
            messages.success(request, f"Cliente '{nome}' excluído com sucesso!")
            return redirect("organizations:cliente_list")
            
        except Exception as e:
            messages.error(request, f"Erro ao excluir cliente: {str(e)}")
    
    return render(request, "organizations/cliente_confirm_delete.html", {"cliente": cliente})


@login_required
def cliente_import(request):
    """Importar clientes via CSV."""
    padarias = get_user_padarias(request.user)
    
    if not padarias.exists():
        messages.error(request, "Você não tem acesso a nenhuma padaria.")
        return redirect("organizations:list")
    
    padaria = padarias.first()
    
    if request.method == "POST":
        csv_file = request.FILES.get("csv_file")
        
        if not csv_file:
            messages.error(request, "Selecione um arquivo CSV.")
            return render(request, "organizations/cliente_import.html")
        
        if not csv_file.name.endswith('.csv'):
            messages.error(request, "O arquivo deve ser um CSV.")
            return render(request, "organizations/cliente_import.html")
        
        try:
            import csv
            import io
            
            decoded_file = csv_file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            
            importados = 0
            erros = 0
            
            for row in reader:
                nome = row.get('nome', '').strip()
                telefone = row.get('telefone', '').strip()
                email = row.get('email', '').strip()
                
                if nome and telefone:
                    try:
                        Cliente.objects.get_or_create(
                            padaria=padaria,
                            telefone=telefone,
                            defaults={
                                'nome': nome,
                                'email': email,
                                'aceita_promocoes': True
                            }
                        )
                        importados += 1
                    except Exception:
                        erros += 1
                else:
                    erros += 1
            
            messages.success(request, f"Importação concluída! {importados} clientes importados, {erros} erros.")
            return redirect("organizations:cliente_list")
            
        except Exception as e:
            messages.error(request, f"Erro ao importar: {str(e)}")
    
    return render(request, "organizations/cliente_import.html")


# =====================================================
# CAMPANHAS WHATSAPP
# =====================================================

@login_required
def campanha_list(request):
    """Lista de campanhas da padaria."""
    padarias = get_user_padarias(request.user)
    
    if not padarias.exists():
        messages.warning(request, "Você não tem acesso a nenhuma padaria.")
        return redirect("organizations:list")
    
    if request.user.is_superuser:
        campanhas = CampanhaWhatsApp.objects.filter(
            padaria__in=padarias
        ).select_related('padaria', 'promocao').order_by('-created_at')
    else:
        padaria = padarias.first()
        campanhas = CampanhaWhatsApp.objects.filter(
            padaria=padaria
        ).select_related('promocao').order_by('-created_at')
    
    context = {
        'campanhas': campanhas,
        'padarias': padarias,
    }
    return render(request, "organizations/campanha_list.html", context)


@login_required
def campanha_create(request):
    """Criar nova campanha."""
    padarias = get_user_padarias(request.user)
    
    if not padarias.exists():
        messages.error(request, "Você não tem acesso a nenhuma padaria.")
        return redirect("organizations:list")
    
    padaria = padarias.first()
    promocoes = Promocao.objects.filter(padaria=padaria, is_active=True)
    clientes_disponiveis = Cliente.objects.filter(
        padaria=padaria,
        is_active=True,
        aceita_promocoes=True
    ).count()
    
    if request.method == "POST":
        nome = request.POST.get("nome", "").strip()
        mensagem = request.POST.get("mensagem", "").strip()
        promocao_id = request.POST.get("promocao")
        imagem = request.FILES.get("imagem")
        delay_minimo = int(request.POST.get("delay_minimo", 15))
        delay_maximo = int(request.POST.get("delay_maximo", 45))
        lote_tamanho = int(request.POST.get("lote_tamanho", 10))
        pausa_entre_lotes = int(request.POST.get("pausa_entre_lotes", 120))
        
        if not nome or not mensagem:
            messages.error(request, "Nome e mensagem são obrigatórios.")
            return render(request, "organizations/campanha_form.html", {
                "padarias": padarias,
                "promocoes": promocoes,
                "clientes_disponiveis": clientes_disponiveis,
            })
        
        try:
            promocao = None
            if promocao_id:
                promocao = Promocao.objects.get(pk=promocao_id)
            
            # Se não enviou imagem nova, mas tem promoção com imagem, usar a da promoção
            if not imagem and promocao and promocao.imagem:
                imagem = promocao.imagem
            
            campanha = CampanhaWhatsApp.objects.create(
                padaria=padaria,
                promocao=promocao,
                nome=nome,
                mensagem=mensagem,
                imagem=imagem,
                delay_minimo=delay_minimo,
                delay_maximo=delay_maximo,
                lote_tamanho=lote_tamanho,
                pausa_entre_lotes=pausa_entre_lotes,
                status='rascunho'
            )
            
            # Adicionar clientes à campanha
            clientes = Cliente.objects.filter(
                padaria=padaria,
                is_active=True,
                aceita_promocoes=True
            )
            
            for cliente in clientes:
                MensagemCampanha.objects.create(
                    campanha=campanha,
                    cliente=cliente,
                    status='pendente'
                )
            
            campanha.total_destinatarios = clientes.count()
            campanha.save()
            
            AuditLog.log(
                action="create",
                entity="CampanhaWhatsApp",
                padaria=padaria,
                actor=request.user,
                entity_id=campanha.id,
                diff={"nome": nome, "destinatarios": clientes.count()}
            )
            
            messages.success(request, f"Campanha '{nome}' criada com {clientes.count()} destinatários!")
            return redirect("organizations:campanha_detail", pk=campanha.pk)
            
        except Exception as e:
            messages.error(request, f"Erro ao criar campanha: {str(e)}")
    
    context = {
        "padarias": padarias,
        "promocoes": promocoes,
        "clientes_disponiveis": clientes_disponiveis,
    }
    return render(request, "organizations/campanha_form.html", context)


@login_required
def campanha_detail(request, pk):
    """Detalhes da campanha."""
    padarias = get_user_padarias(request.user)
    campanha = get_object_or_404(CampanhaWhatsApp, pk=pk)
    
    if not request.user.is_superuser and campanha.padaria not in padarias:
        messages.error(request, "Você não tem acesso a esta campanha.")
        return redirect("organizations:campanha_list")
    
    mensagens = campanha.mensagens.select_related('cliente').order_by('id')[:50]
    
    context = {
        'campanha': campanha,
        'mensagens': mensagens,
        'total_mensagens': campanha.mensagens.count(),
    }
    return render(request, "organizations/campanha_detail.html", context)


@login_required
def campanha_iniciar(request, pk):
    """Iniciar envio da campanha."""
    padarias = get_user_padarias(request.user)
    campanha = get_object_or_404(CampanhaWhatsApp, pk=pk)
    
    if not request.user.is_superuser and campanha.padaria not in padarias:
        messages.error(request, "Você não tem acesso a esta campanha.")
        return redirect("organizations:campanha_list")
    
    if request.method != "POST":
        return redirect("organizations:campanha_detail", pk=pk)
    
    if campanha.status not in ['rascunho', 'pausada']:
        messages.error(request, "Esta campanha não pode ser iniciada.")
        return redirect("organizations:campanha_detail", pk=pk)
    
    from .campaign_service import CampaignService
    
    service = CampaignService(campanha)
    
    # Verificar conexão WhatsApp
    conectado, msg = service.verificar_conexao()
    if not conectado:
        messages.error(request, f"WhatsApp não conectado: {msg}")
        return redirect("organizations:campanha_detail", pk=pk)
    
    # Iniciar envio em background
    service.executar_campanha(async_mode=True)
    
    AuditLog.log(
        action="campanha_iniciada",
        entity="CampanhaWhatsApp",
        padaria=campanha.padaria,
        actor=request.user,
        entity_id=campanha.id,
        diff={"status": "enviando"}
    )
    
    messages.success(request, "Campanha iniciada! O envio está sendo processado em segundo plano.")
    return redirect("organizations:campanha_detail", pk=pk)


@login_required
def campanha_pausar(request, pk):
    """Pausar campanha em andamento."""
    padarias = get_user_padarias(request.user)
    campanha = get_object_or_404(CampanhaWhatsApp, pk=pk)
    
    if not request.user.is_superuser and campanha.padaria not in padarias:
        messages.error(request, "Você não tem acesso a esta campanha.")
        return redirect("organizations:campanha_list")
    
    if request.method != "POST":
        return redirect("organizations:campanha_detail", pk=pk)
    
    if campanha.status != 'enviando':
        messages.error(request, "Esta campanha não está em envio.")
        return redirect("organizations:campanha_detail", pk=pk)
    
    from .campaign_service import pausar_campanha
    
    if pausar_campanha(campanha.id):
        messages.success(request, "Solicitação de pausa enviada. A campanha será pausada após a mensagem atual.")
    else:
        messages.warning(request, "Não foi possível pausar. A campanha pode já ter sido concluída.")
    
    return redirect("organizations:campanha_detail", pk=pk)


@login_required
def campanha_status_ajax(request, pk):
    """Retorna status atualizado da campanha (para atualização em tempo real)."""
    campanha = get_object_or_404(CampanhaWhatsApp, pk=pk)
    
    return JsonResponse({
        'status': campanha.status,
        'status_display': campanha.get_status_display(),
        'enviados': campanha.enviados,
        'falhas': campanha.falhas,
        'total': campanha.total_destinatarios,
        'progresso': campanha.get_progresso(),
    })


@login_required
def campanha_delete(request, pk):
    """Deletar campanha."""
    padarias = get_user_padarias(request.user)
    campanha = get_object_or_404(CampanhaWhatsApp, pk=pk)
    
    if not request.user.is_superuser and campanha.padaria not in padarias:
        messages.error(request, "Você não tem acesso a esta campanha.")
        return redirect("organizations:campanha_list")
    
    if campanha.status == 'enviando':
        messages.error(request, "Não é possível excluir uma campanha em andamento. Pause primeiro.")
        return redirect("organizations:campanha_detail", pk=pk)
    
    if request.method == "POST":
        nome = campanha.nome
        
        try:
            AuditLog.log(
                action="delete",
                entity="CampanhaWhatsApp",
                padaria=campanha.padaria,
                actor=request.user,
                entity_id=campanha.id,
                diff={"nome": nome}
            )
            
            campanha.delete()
            messages.success(request, f"Campanha '{nome}' excluída com sucesso!")
            return redirect("organizations:campanha_list")
            
        except Exception as e:
            messages.error(request, f"Erro ao excluir campanha: {str(e)}")
    
    return render(request, "organizations/campanha_confirm_delete.html", {"campanha": campanha})


@login_required
def campanha_criar_de_promocao(request, promocao_pk):
    """Criar campanha rapidamente a partir de uma promoção."""
    padarias = get_user_padarias(request.user)
    promocao = get_object_or_404(Promocao, pk=promocao_pk)
    
    if not request.user.is_superuser and promocao.padaria not in padarias:
        messages.error(request, "Você não tem acesso a esta promoção.")
        return redirect("organizations:promocao_list")
    
    from .campaign_service import criar_campanha_promocao
    
    try:
        campanha = criar_campanha_promocao(promocao)
        
        AuditLog.log(
            action="create",
            entity="CampanhaWhatsApp",
            padaria=promocao.padaria,
            actor=request.user,
            entity_id=campanha.id,
            diff={"nome": campanha.nome, "promocao_id": promocao.id}
        )
        
        messages.success(request, f"Campanha criada com {campanha.total_destinatarios} destinatários!")
        return redirect("organizations:campanha_detail", pk=campanha.pk)
        
    except Exception as e:
        messages.error(request, f"Erro ao criar campanha: {str(e)}")
        return redirect("organizations:promocao_list")



def _configure_webhook(api_url, headers, instance_name):
    """
    Configura webhook automaticamente para a instância.
    
    Configura:
    - URL: https://n8n.newcouros.com.br/webhook/vendedor_pandia
    - Enabled: True
    - Evento: MESSAGES_UPSERT
    """
    try:
        webhook_url = f"{api_url}/webhook/set/{instance_name}"
        webhook_payload = {
            "url": "https://n8n.newcouros.com.br/webhook/vendedor_pandia",
            "enabled": True,
            "webhookByEvents": False,
            "webhookBase64": False,
            "events": [
                "MESSAGES_UPSERT"
            ]
        }
        
        response = requests.post(
            webhook_url,
            headers=headers,
            json=webhook_payload,
            timeout=10
        )
        
        # Log para debug (opcional)
        import logging
        logger = logging.getLogger(__name__)
        
        if response.status_code in [200, 201]:
            logger.info(f"Webhook configurado com sucesso para {instance_name}")
        else:
            logger.warning(f"Erro ao configurar webhook para {instance_name}: {response.text}")
            
    except Exception as e:
        # Não interromper o fluxo se o webhook falhar
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Exceção ao configurar webhook para {instance_name}: {str(e)}")


@login_required
def report_list(request):
    """Redireciona para o relatório da primeira padaria encontrada."""
    padarias = get_user_padarias(request.user)
    if padarias.exists():
        return redirect('organizations:report_detail', slug=padarias.first().slug)
    
    messages.warning(request, "Você precisa ter uma padaria para acessar relatórios.")
    return redirect('organizations:list')


@login_required
def report_detail(request, slug):
    """
    Relatórios detalhados da padaria (Vendas, Produtos, Faturamento).
    Busca dados diretamente da API do Mercado Pago para incluir vendas externas.
    """
    from django.apps import apps
    import json
    from django.core.serializers.json import DjangoJSONEncoder
    from datetime import datetime, timedelta
    from collections import defaultdict
    from payments.services.mercadopago_service import get_mp_service, MercadoPagoAPIError

    padarias = get_user_padarias(request.user)
    organization = get_object_or_404(Padaria, slug=slug)
    
    # Verificar acesso
    if not request.user.is_superuser and organization not in padarias:
        messages.error(request, "Você não tem acesso a esta padaria.")
        return redirect("organizations:list")

    # Inicializar dados vazios
    all_payments = []
    prev_revenue = 0
    prev_count = 0
    
    # Filtros de Data
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Configurar datas padrão se não fornecidas (últimos 30 dias)
    if not start_date_str:
        start_date_dt = datetime.now() - timedelta(days=30)
    else:
        start_date_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
        
    if not end_date_str:
        end_date_dt = datetime.now()
    else:
        end_date_dt = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)

    # Buscar dados do Mercado Pago
    try:
        mp_service = get_mp_service(organization)
        
        # Filtros atuais
        filters = {
            "sort": "date_created",
            "criteria": "desc",
            "range": "date_created",
            "begin_date": start_date_dt.strftime("%Y-%m-%dT%H:%M:%S.000-03:00"),
            "end_date": end_date_dt.strftime("%Y-%m-%dT%H:%M:%S.000-03:00"),
            "limit": 1000
        }
        
        search_result = mp_service.search_payments(**filters)
        all_payments = search_result.get("results", [])

        # Buscar período anterior para comparação (Growth)
        delta = end_date_dt - start_date_dt
        prev_end_date = start_date_dt - timedelta(seconds=1)
        prev_start_date = prev_end_date - delta
        
        filters_prev = {
            "range": "date_created",
            "begin_date": prev_start_date.strftime("%Y-%m-%dT%H:%M:%S.000-03:00"),
            "end_date": prev_end_date.strftime("%Y-%m-%dT%H:%M:%S.000-03:00"),
            "limit": 1000,
            "status": "approved" # Buscar apenas aprovados para otimizar métricas financeiras
        }
        prev_search = mp_service.search_payments(**filters_prev)
        prev_payments = prev_search.get("results", [])
        
        prev_revenue = sum(float(p.get('transaction_amount', 0)) for p in prev_payments)
        prev_count = len(prev_payments)
        
    except ValueError:
        pass 
    except Exception as e:
        messages.warning(request, f"Erro ao sincronizar com Mercado Pago: {str(e)}")
        pass

    # Processamento dos Dados Atuais
    approved_payments = [p for p in all_payments if p.get('status') == 'approved']
    
    # 1. Métricas Principais & Growth
    total_revenue = sum(float(p.get('transaction_amount', 0)) for p in approved_payments)
    total_sales_count = len(approved_payments)
    avg_ticket = total_revenue / total_sales_count if total_sales_count > 0 else 0
    
    # Calcular crescimentos
    def calc_growth(current, previous):
        if previous == 0:
            return 100 if current > 0 else 0
        return ((current - previous) / previous) * 100

    revenue_growth = calc_growth(total_revenue, prev_revenue)
    sales_growth = calc_growth(total_sales_count, prev_count)
    
    # Taxa de Reembolso/Cancelamento
    refunded_count = len([p for p in all_payments if p.get('status') in ['refunded', 'cancelled', 'rejected']])
    total_tx = len(all_payments)
    refund_rate = (refunded_count / total_tx * 100) if total_tx > 0 else 0

    # 2. Vendas por Data (Gráfico de Linha)
    sales_by_date = defaultdict(lambda: {'revenue': 0.0, 'count': 0})
    for p in approved_payments:
        date_str = p.get('date_created', '')[:10]
        sales_by_date[date_str]['revenue'] += float(p.get('transaction_amount', 0))

    sorted_dates = sorted(sales_by_date.keys())
    chart_labels = []
    chart_data = []
    
    for d in sorted_dates:
        try:
            date_obj = datetime.strptime(d, '%Y-%m-%d')
            chart_labels.append(date_obj.strftime('%d/%m'))
        except:
            chart_labels.append(d)
        chart_data.append(sales_by_date[d]['revenue'])

    # 3. Distribuição de Status (Gráfico Donut)
    # Dicionário de tradução de status
    status_traducao = {
        'accredited': 'Aprovado',
        'approved': 'Aprovado',
        'pending': 'Pendente',
        'in_process': 'Em Processamento',
        'rejected': 'Rejeitado',
        'cancelled': 'Cancelado',
        'refunded': 'Reembolsado',
        'charged_back': 'Estornado',
        'cc_rejected_high_risk': 'Recusado - Alto Risco',
        'cc_rejected_call_for_authorize': 'Recusado - Autorização Necessária',
        'cc_rejected_insufficient_amount': 'Recusado - Saldo Insuficiente',
        'cc_rejected_bad_filled_card_number': 'Recusado - Número do Cartão Inválido',
        'cc_rejected_bad_filled_date': 'Recusado - Data Inválida',
        'cc_rejected_bad_filled_security_code': 'Recusado - CVV Inválido',
        'cc_rejected_blacklist': 'Recusado - Bloqueado',
        'cc_rejected_duplicated_payment': 'Recusado - Pagamento Duplicado',
        'cc_rejected_other_reason': 'Recusado - Outros',
        'bpp_refunded': 'Reembolsado',
        'pending_waiting_payment': 'Aguardando Pagamento',
        'pending_waiting_transfer': 'Aguardando Transferência',
    }
    
    status_dist = defaultdict(int)
    for p in all_payments:
        status_original = p.get('status_detail') or p.get('status')
        status_traduzido = status_traducao.get(status_original, status_original)
        status_dist[status_traduzido] += 1
    
    status_labels = list(status_dist.keys())
    status_data = list(status_dist.values())

    # 4. Produtos Vendidos
    product_stats = defaultdict(lambda: {'qty': 0, 'revenue': 0.0})
    for p in approved_payments:
        items = p.get('additional_info', {}).get('items', [])
        if items:
            for item in items:
                title = item.get('title', 'Produto sem nome')
                qty = float(item.get('quantity', 1))
                price = float(item.get('unit_price', 0))
                product_stats[title]['qty'] += int(qty)
                product_stats[title]['revenue'] += (price * qty)
        else:
            desc = p.get('description') or p.get('reason') or "Venda Diversa"
            amount = float(p.get('transaction_amount', 0))
            product_stats[desc]['qty'] += 1
            product_stats[desc]['revenue'] += amount

    products_sold = [
        {'description': k, 'qty': v['qty'], 'revenue': v['revenue']} 
        for k, v in product_stats.items()
    ]
    products_sold.sort(key=lambda x: x['revenue'], reverse=True)

    # 5. Métodos de Pagamento (Gráfico Pizza + Tabela)
    # Dicionário de tradução de métodos de pagamento
    payment_method_traducao = {
        'pix': 'PIX',
        'credit_card': 'Cartão de Crédito',
        'debit_card': 'Cartão de Débito',
        'account_money': 'Saldo Mercado Pago',
        'master': 'Mastercard',
        'visa': 'Visa',
        'amex': 'American Express',
        'elo': 'Elo',
        'hipercard': 'Hipercard',
        'bolbradesco': 'Boleto Bradesco',
        'ticket': 'Boleto',
        'bank_transfer': 'Transferência Bancária',
        'atm': 'Caixa Eletrônico',
        'prepaid_card': 'Cartão Pré-pago',
        'digital_wallet': 'Carteira Digital',
        'voucher_card': 'Voucher',
        'crypto_transfer': 'Criptomoeda',
        'digital_currency': 'Moeda Digital',
    }
    
    payment_methods = defaultdict(lambda: {'count': 0, 'revenue': 0.0})
    for p in approved_payments:
        method_original = p.get('payment_method_id', 'outros')
        method_traduzido = payment_method_traducao.get(method_original, method_original.capitalize())
        payment_methods[method_traduzido]['count'] += 1
        payment_methods[method_traduzido]['revenue'] += float(p.get('transaction_amount', 0))
    
    payment_methods_list = [
        {'method': k, 'count': v['count'], 'revenue': v['revenue']}
        for k, v in payment_methods.items()
    ]
    payment_methods_list.sort(key=lambda x: x['revenue'], reverse=True)
    
    pm_labels = [pm['method'] for pm in payment_methods_list]
    pm_data = [pm['revenue'] for pm in payment_methods_list]

    # 6. Transações Recentes
    recent_transactions = []
    for p in all_payments[:50]:
        created_at = p.get('date_created')
        try:
            created_at_dt = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%S.000-03:00")
        except:
            created_at_dt = created_at
            
        recent_transactions.append({
            'created_at': created_at_dt,
            'status': p.get('status'),
            'get_status_display': p.get('status_detail') or p.get('status'),
            'amount': float(p.get('transaction_amount', 0)),
            'description': p.get('description') or "Pagamento Mercado Pago"
        })

    context = {
        'organization': organization,
        'total_revenue': total_revenue,
        'total_sales_count': total_sales_count,
        'avg_ticket': avg_ticket,
        'revenue_growth': revenue_growth,
        'sales_growth': sales_growth,
        'refund_rate': refund_rate,
        # Charts Data
        'chart_labels': json.dumps(chart_labels, cls=DjangoJSONEncoder),
        'chart_data': json.dumps(chart_data, cls=DjangoJSONEncoder),
        'status_labels': json.dumps(status_labels, cls=DjangoJSONEncoder),
        'status_data': json.dumps(status_data, cls=DjangoJSONEncoder),
        'pm_labels': json.dumps(pm_labels, cls=DjangoJSONEncoder),
        'pm_data': json.dumps(pm_data, cls=DjangoJSONEncoder),
        # Lists
        'products_sold': products_sold,
        'payment_methods': payment_methods_list,
        'recent_transactions': recent_transactions,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'mp_connected': bool(all_payments) # Flag para saber se tem dados
    }

    return render(request, "organizations/reports.html", context)